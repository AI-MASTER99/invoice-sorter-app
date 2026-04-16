# v1776278049645 - supabase migration
import asyncio
import base64
import io
import json
import os
import re
import threading
import time
import uuid
from pathlib import Path
from dotenv import load_dotenv
load_dotenv(dotenv_path=Path(__file__).parent / ".env")
from datetime import date, datetime
from typing import Any
import anthropic
import httpx
import openpyxl
from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from starlette.middleware.sessions import SessionMiddleware
from supabase import create_client, Client

BASE_DIR = Path(__file__).parent.parent / "invoiceflow"
_ON_VERCEL = os.environ.get("VERCEL") == "1"
if _ON_VERCEL:
    _DATA_ROOT = Path("/tmp/invoice-sorter")
else:
    _DATA_ROOT = BASE_DIR
UPLOADS_DIR = _DATA_ROOT / "uploads"
OUTPUT_DIR = _DATA_ROOT / "output"
USERS_FILE = _DATA_ROOT / "users.json"
for d in (UPLOADS_DIR, OUTPUT_DIR):
    d.mkdir(parents=True, exist_ok=True)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
STORAGE_BUCKET = "invoice-exports"
_supabase: Client | None = None

def get_supabase():
    global _supabase
    if _supabase is None and SUPABASE_URL and SUPABASE_KEY:
        try:
            _supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        except Exception as e:
            print(f"[supabase] init failed: {e}")
            _supabase = None
    return _supabase

def load_users():
    default_user = os.environ.get("APP_USERNAME", "admin")
    default_pw = os.environ.get("APP_PASSWORD", "changeme")
    return {default_user: {"password": default_pw, "role": "admin"}}

def verify_password(plain: str, hashed: str) -> bool:
    return plain == hashed

def require_auth(request: Request) -> str:
    user = request.session.get("user")
    if not user:
        import hmac, hashlib, time
        token = request.cookies.get("auth_token","")
        if token and len(token.split(":")) == 3:
            u, ts, sig = token.split(":")
            secret = os.environ.get("SECRET_KEY","dev-secret-change-me")
            expected = hmac.new(secret.encode(), f"{u}:{ts}".encode(), hashlib.sha256).hexdigest()
            if hmac.compare_digest(sig, expected) and int(time.time()) - int(ts) < 43200:
                user = u
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

_lock = threading.Lock()
jobs = {}
processed_today = 0
_today_date = date.today().isoformat()

def _reset_daily():
    global processed_today, _today_date
    today = date.today().isoformat()
    if today != _today_date:
        with _lock:
            processed_today = 0
            _today_date = today

def load_memory():
    sb = get_supabase()
    if not sb: return {}
    try:
        res = sb.table("product_memory").select("*").execute()
        result = {}
        for row in (res.data or []):
            key = row.get("key")
            if not key: continue
            result[key] = {
                "code": row.get("code",""),
                "description": row.get("description",""),
                "confirmed": row.get("confirmed", False),
                "tariff": row.get("tariff") or {},
            }
        return result
    except Exception as e:
        print(f"[load_memory] error: {e}")
        return {}

def save_memory(data):
    sb = get_supabase()
    if not sb: return
    try:
        records = []
        for key, val in data.items():
            records.append({
                "key": key,
                "code": val.get("code",""),
                "description": val.get("description",""),
                "confirmed": bool(val.get("confirmed", False)),
                "tariff": val.get("tariff") or {},
            })
        if records:
            sb.table("product_memory").upsert(records).execute()
    except Exception as e:
        print(f"[save_memory] error: {e}")

def save_processed_invoice(inv):
    sb = get_supabase()
    if not sb: return
    try:
        sb.table("processed_invoices").upsert({
            "id": inv["id"],
            "supplier": inv.get("supplier",""),
            "filename": inv.get("filename",""),
            "invoice_date": inv.get("date") or datetime.now().isoformat(),
            "value": inv.get("value",""),
            "status": inv.get("status",""),
            "full_xlsx_path": inv.get("full_xlsx_path",""),
            "raw_xlsx_path": inv.get("raw_xlsx_path",""),
            "rows": inv.get("rows") or [],
            "tariff_data": inv.get("tariff_data") or {},
            "user_name": inv.get("user_name",""),
        }).execute()
    except Exception as e:
        print(f"[save_invoice] error: {e}")

def list_processed_invoices():
    sb = get_supabase()
    if not sb: return []
    try:
        res = sb.table("processed_invoices").select("id,supplier,filename,invoice_date,value,status").order("invoice_date", desc=True).execute()
        return [
            {"id":r["id"],"supplier":r.get("supplier",""),"filename":r.get("filename",""),
             "date":r.get("invoice_date",""),"value":r.get("value",""),"status":r.get("status","")}
            for r in (res.data or [])
        ]
    except Exception as e:
        print(f"[list_invoices] error: {e}")
        return []

def get_processed_invoice(invoice_id):
    sb = get_supabase()
    if not sb: return None
    try:
        res = sb.table("processed_invoices").select("*").eq("id", invoice_id).limit(1).execute()
        rows = res.data or []
        return rows[0] if rows else None
    except Exception as e:
        print(f"[get_invoice] error: {e}")
        return None

def upload_to_storage(local_path, remote_path):
    sb = get_supabase()
    if not sb: return ""
    try:
        data = local_path.read_bytes()
        sb.storage.from_(STORAGE_BUCKET).upload(
            path=remote_path, file=data,
            file_options={"content-type":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet","upsert":"true"}
        )
        return remote_path
    except Exception as e:
        print(f"[upload_storage] error: {e}")
        return ""

def download_from_storage(remote_path):
    sb = get_supabase()
    if not sb: return None
    try:
        return sb.storage.from_(STORAGE_BUCKET).download(remote_path)
    except Exception as e:
        print(f"[download_storage] error: {e}")
        return None

async def lookup_tariff(commodity_code):
    code = re.sub(r"\\D", "", commodity_code)
    if not code: return {}
    url = f"https://www.trade-tariff.service.gov.uk/api/v2/commodities/{code}"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(url)
            if r.status_code != 200: return {}
            data = r.json()
            duty = ""
            vat = ""
            measures = data.get("included", [])
            for m in measures:
                if not isinstance(m, dict): continue
                mtype = m.get("attributes", {}).get("measure_type_description", "")
                duty_expr = m.get("attributes", {}).get("duty_expression", {})
                formatted = duty_expr.get("formatted_base", "") if isinstance(duty_expr, dict) else ""
                if "Third country duty" in mtype and not duty:
                    duty = formatted
                if "VAT" in mtype and not vat:
                    vat = formatted
            desc = data.get("data", {}).get("attributes", {}).get("description", "")
            return {"description": desc, "duty": duty or "N/A", "vat": vat or "20%"}
    except Exception:
        return {}

PROMPT_EXTRACT = """You are an invoice data extraction expert. Extract all line items from the invoice into a TAB-SEPARATED table.
Columns: Invoice, Comm./imp. cod, Description of Goods, Origin, Country, Number of Packages, Gross Weight (KG), Net Weight (KG), Value
Rules: Extract only what is explicitly on the invoice. Leave blank if missing. Output TSV only, no extra text."""

PROMPT_VERIFY = "Second independent extraction. " + PROMPT_EXTRACT
COLUMNS = ["Invoice","Comm./imp. cod","Description of Goods","Origin","Country","Number of Packages","Gross Weight (KG)","Net Weight (KG)","Value"]

def parse_tsv(tsv):
    lines = [l for l in tsv.strip().splitlines() if l.strip()]
    if not lines: return []
    header_line = 0
    for i, line in enumerate(lines):
        if "Invoice" in line or "Comm" in line:
            header_line = i
            break
    headers = [h.strip() for h in lines[header_line].split("\\t")]
    rows = []
    for line in lines[header_line + 1:]:
        parts = line.split("\\t")
        while len(parts) < len(headers):
            parts.append("")
        row = {headers[i]: parts[i].strip() for i in range(len(headers))}
        rows.append(row)
    return rows

def normalise_row(row):
    mapping = {"invoice":"Invoice","comm./imp. cod":"Comm./imp. cod","description of goods":"Description of Goods","origin":"Origin","country":"Country","number of packages":"Number of Packages","gross weight (kg)":"Gross Weight (KG)","net weight (kg)":"Net Weight (KG)","value":"Value"}
    return {mapping.get(k.lower().strip(), k): v for k, v in row.items()}

def rows_match(a, b):
    if len(a) != len(b): return False
    for ra, rb in zip(a, b):
        for field in ("Comm./imp. cod","Value","Gross Weight (KG)","Net Weight (KG)"):
            va = re.sub(r"[^\\d.]", "", ra.get(field, "") or "")
            vb = re.sub(r"[^\\d.]", "", rb.get(field, "") or "")
            if va and vb and va != vb:
                return False
    return True

def extract_value_number(val_str):
    if not val_str: return None
    cleaned = re.sub(r"[^\\d.\\-]", "", val_str)
    try:
        return float(cleaned)
    except ValueError:
        return None

_FILL_HEADER = PatternFill("solid", fgColor="1F3864")
_FILL_ALT = PatternFill("solid", fgColor="DCE6F1")
_FILL_TOTALS = PatternFill("solid", fgColor="2E75B6")
_FONT_HDR = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_FONT_CELL = Font(name="Calibri", color="000000", size=10)
_FONT_TOTALS = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_COL_CFG = [("Invoice",12,"General","left"),("Comm./imp. cod",18,"@","left"),("Description of Goods",42,"General","left"),("Origin",8,"General","center"),("Country",12,"General","left"),("Number of Packages",8,"#,##0.00","right"),("Gross Weight (KG)",18,"#,##0.00","right"),("Net Weight (KG)",16,"#,##0.00","right"),("Value",14,"\u20ac#,##0.00","right")]

def build_excel(rows, tariff_data, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = sheet_title
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    for col_idx, (col_name, width, fmt, align) in enumerate(_COL_CFG, start=1):
        c = ws.cell(row=2, column=col_idx, value=col_name)
        c.font = _FONT_HDR
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"
    data_start = 3
    for row_idx, row in enumerate(rows, start=data_start):
        fill = _FILL_ALT if (row_idx % 2 == 0) else None
        for col_idx, (col_name, width, fmt, align) in enumerate(_COL_CFG, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            raw = row.get(col_name, "") or ""
            if fill:
                c.fill = fill
            c.font = _FONT_CELL
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.number_format = fmt
            if col_name in ("Value","Gross Weight (KG)","Net Weight (KG)","Number of Packages"):
                num = extract_value_number(str(raw))
                c.value = num if num is not None else raw
            else:
                c.value = str(raw) if raw else ""
    data_end = data_start + len(rows) - 1
    total_row = data_end + 1
    ws.merge_cells(f"A{total_row}:E{total_row}")
    tc = ws.cell(row=total_row, column=1, value="TOTALS")
    tc.font = _FONT_TOTALS
    tc.fill = _FILL_TOTALS
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx, (col_name, width, fmt, align) in enumerate(_COL_CFG, start=1):
        if col_idx < 6:
            ws.cell(row=total_row, column=col_idx).fill = _FILL_TOTALS
            continue
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx)
        c.fill = _FILL_TOTALS
        c.font = _FONT_TOTALS
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = fmt
        if col_name in ("Gross Weight (KG)","Net Weight (KG)","Value"):
            c.value = f"=ROUND(SUM({col_letter}{data_start}:{col_letter}{data_end}),2)"
        else:
            c.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
    for col_idx, (col_name, width, fmt, align) in enumerate(_COL_CFG, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

async def run_extraction(client, file_bytes, mime, prompt):
    b64 = base64.standard_b64encode(file_bytes).decode()
    if mime == "application/pdf":
        doc_block = {"type":"document","source":{"type":"base64","media_type":mime,"data":b64},"cache_control":{"type":"ephemeral"}}
    elif mime.startswith("image/"):
        doc_block = {"type":"image","source":{"type":"base64","media_type":mime,"data":b64},"cache_control":{"type":"ephemeral"}}
    else:
        doc_block = {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64},"cache_control":{"type":"ephemeral"}}
    message = await client.messages.create(model="claude-opus-4-6",max_tokens=4096,messages=[{"role":"user","content":[doc_block,{"type":"text","text":prompt}]}])
    return message.content[0].text

async def _process_invoice(job_id, file_path, original_name, mime):
    global processed_today
    def update(progress, step):
        with _lock:
            if job_id in jobs:
                jobs[job_id]["progress"] = progress
                jobs[job_id]["step"] = step
    invoice_id = str(uuid.uuid4())
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    client = anthropic.AsyncAnthropic(api_key=api_key)
    try:
        file_bytes = file_path.read_bytes()
        update(10, "Extracting data (Run A)...")
        raw_a = await run_extraction(client, file_bytes, mime, PROMPT_EXTRACT)
        update(35, "Verifying (Run B)...")
        raw_b = await run_extraction(client, file_bytes, mime, PROMPT_VERIFY)
        rows_a = [normalise_row(r) for r in parse_tsv(raw_a)]
        rows_b = [normalise_row(r) for r in parse_tsv(raw_b)]
        update(60, "Cross-checking...")
        verified = rows_match(rows_a, rows_b)
        final_rows = rows_a
        attempts = 1
        while not verified and attempts < 3:
            raw_retry = await run_extraction(client, file_bytes, mime, PROMPT_VERIFY)
            rows_retry = [normalise_row(r) for r in parse_tsv(raw_retry)]
            verified = rows_match(rows_a, rows_retry)
            if verified:
                final_rows = rows_a
            attempts += 1
        status = "verified" if verified else "subcode_needed"
        update(80, "Looking up tariff codes...")
        tariff_data = {}
        seen_codes = set()
        for row in final_rows:
            code = row.get("Comm./imp. cod","").strip()
            if code and code not in seen_codes:
                seen_codes.add(code)
                info = await lookup_tariff(code)
                tariff_data[code] = info
        memory = load_memory()
        for row in final_rows:
            code = row.get("Comm./imp. cod","").strip()
            desc = row.get("Description of Goods","").strip()
            if code and desc:
                key = f"{code}::{desc}"
                tariff_info = tariff_data.get(code, {})
                if key not in memory:
                    memory[key] = {"code":code,"description":desc,"confirmed":verified,"tariff":tariff_info}
                else:
                    if not memory[key].get("tariff") and tariff_info:
                        memory[key]["tariff"] = tariff_info
                    if verified:
                        memory[key]["confirmed"] = True
        save_memory(memory)
        update(95, "Generating Excel files...")
        stem = Path(original_name).stem
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_stem = re.sub(r"[^\\w\\-]", "_", stem)
        local_full = OUTPUT_DIR / f"{safe_stem}_{ts}_full.xlsx"
        local_raw = OUTPUT_DIR / f"{safe_stem}_{ts}_raw.xlsx"
        full_bytes = build_excel(final_rows, tariff_data, "Invoice Data")
        raw_bytes = build_excel(final_rows, None, "Raw Extraction")
        local_full.write_bytes(full_bytes)
        local_raw.write_bytes(raw_bytes)
        remote_full = f"{invoice_id}/{safe_stem}_{ts}_full.xlsx"
        remote_raw = f"{invoice_id}/{safe_stem}_{ts}_raw.xlsx"
        upload_to_storage(local_full, remote_full)
        upload_to_storage(local_raw, remote_raw)
        supplier = final_rows[0].get("Invoice", original_name) if final_rows else original_name
        total_value = 0.0
        currency = "\u20ac"
        for row in final_rows:
            v = row.get("Value","") or ""
            num = extract_value_number(v)
            if num:
                total_value += num
            if "\u00a3" in v:
                currency = "\u00a3"
            elif "$" in v:
                currency = "$"
        with _lock:
            processed_today += 1
        invoice_record = {
            "id":invoice_id,"supplier":supplier,"filename":original_name,
            "date":datetime.now().isoformat(),"value":f"{currency}{total_value:,.2f}",
            "status":status,"full_xlsx_path":remote_full,"raw_xlsx_path":remote_raw,
            "rows":final_rows,"tariff_data":tariff_data
        }
        save_processed_invoice(invoice_record)
        jobs[job_id]["status"] = "done"
        jobs[job_id]["invoice_id"] = invoice_id
        jobs[job_id]["progress"] = 100
        jobs[job_id]["step"] = "Complete"
    except Exception as exc:
        with _lock:
            if job_id in jobs:
                jobs[job_id]["status"] = "failed"
                jobs[job_id]["step"] = f"Error: {exc}"
                jobs[job_id]["progress"] = 0
        raise

def _run_pipeline(job_id, file_path, original_name, mime):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_process_invoice(job_id, file_path, original_name, mime))
    finally:
        loop.close()

app = FastAPI(title="Invoice Sorter")
app.add_middleware(SessionMiddleware,secret_key=os.environ.get("SECRET_KEY","dev-secret-change-me"),session_cookie="is_session",max_age=60*60*12,https_only=True,same_site="none")
app.add_middleware(CORSMiddleware,allow_origins=["*"],allow_methods=["*"],allow_headers=["*"],allow_credentials=True)
MIME_MAP = {".pdf":"application/pdf",".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",".docx":"application/vnd.openxmlformats-officedocument.wordprocessingml.document"}

@app.get("/login", response_class=HTMLResponse)
async def login_page():
    login_html = BASE_DIR / "static" / "login.html"
    return HTMLResponse(content=login_html.read_text(encoding="utf-8"))

@app.post("/api/login")
async def api_login(request: Request, body: dict = {}):
    username = (body.get("username") or "").strip().lower()
    password = body.get("password") or ""
    users = load_users()
    entry = users.get(username)
    if not entry or not verify_password(password, entry["password"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    request.session["user"] = username
    request.session["role"] = entry.get("role","user")
    import hmac, hashlib, time
    secret = os.environ.get("SECRET_KEY","dev-secret-change-me")
    ts = str(int(time.time()))
    sig = hmac.new(secret.encode(), f"{username}:{ts}".encode(), hashlib.sha256).hexdigest()
    token = f"{username}:{ts}:{sig}"
    from fastapi.responses import JSONResponse
    resp = JSONResponse({"ok":True,"user":username,"role":entry.get("role","user")})
    resp.set_cookie("auth_token", token, max_age=60*60*12, httponly=False, samesite="none", secure=True)
    return resp

@app.post("/api/logout")
async def api_logout(request: Request):
    request.session.clear()
    return {"ok":True}

@app.get("/api/me")
async def api_me(request: Request):
    user = request.session.get("user")
    if not user:
        import hmac, hashlib, time
        token = request.cookies.get("auth_token","")
        if token and len(token.split(":")) == 3:
            u, ts, sig = token.split(":")
            secret = os.environ.get("SECRET_KEY","dev-secret-change-me")
            expected = hmac.new(secret.encode(), f"{u}:{ts}".encode(), hashlib.sha256).hexdigest()
            if hmac.compare_digest(sig, expected) and int(time.time()) - int(ts) < 43200:
                user = u
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return {"user":user,"role":request.session.get("role","user")}

@app.post("/upload")
async def upload_invoice(file: UploadFile = File(...), _: str = Depends(require_auth)):
    _reset_daily()
    ext = Path(file.filename).suffix.lower()
    mime = MIME_MAP.get(ext)
    if not mime:
        raise HTTPException(400, f"Unsupported file type: {ext}")
    job_id = str(uuid.uuid4())
    safe_name = re.sub(r"[^\\w\\-.]","_",file.filename)
    dest = UPLOADS_DIR / f"{job_id}_{safe_name}"
    content = await file.read()
    dest.write_bytes(content)
    with _lock:
        jobs[job_id] = {"id":job_id,"filename":file.filename,"status":"running","progress":0,"step":"Queued...","created_at":datetime.now().isoformat(),"invoice_id":None,"error":None}
    t = threading.Thread(target=_run_pipeline,args=(job_id,dest,file.filename,mime),daemon=True)
    t.start()
    return {"job_id":job_id}

@app.get("/jobs")
def list_jobs(_: str = Depends(require_auth)):
    _reset_daily()
    with _lock:
        return list(jobs.values())

@app.get("/stats")
def get_stats(_: str = Depends(require_auth)):
    _reset_daily()
    memory = load_memory()
    pending = sum(1 for v in memory.values() if not v.get("confirmed",True))
    all_invoices = list_processed_invoices()
    total = len(all_invoices)
    verified = sum(1 for v in all_invoices if v["status"]=="verified")
    rate = round(verified/total*100) if total else 0
    with _lock:
        return {"processed_today":processed_today,"verification_rate":rate,"memory_count":len(memory),"memory_pending":pending}

@app.get("/invoices")
def list_invoices(_: str = Depends(require_auth)):
    return list_processed_invoices()

@app.get("/invoices/{invoice_id}/export/full")
def export_full(invoice_id: str, _: str = Depends(require_auth)):
    inv = get_processed_invoice(invoice_id)
    if not inv:
        raise HTTPException(404,"Invoice not found")
    remote = inv.get("full_xlsx_path","")
    if not remote:
        raise HTTPException(404,"Excel file not found")
    data = download_from_storage(remote)
    if not data:
        raise HTTPException(404,"Excel file not found in storage")
    from fastapi.responses import Response
    filename = Path(remote).name
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@app.get("/invoices/{invoice_id}/export/raw")
def export_raw(invoice_id: str, _: str = Depends(require_auth)):
    inv = get_processed_invoice(invoice_id)
    if not inv:
        raise HTTPException(404,"Invoice not found")
    remote = inv.get("raw_xlsx_path","")
    if not remote:
        raise HTTPException(404,"Excel file not found")
    data = download_from_storage(remote)
    if not data:
        raise HTTPException(404,"Excel file not found in storage")
    from fastapi.responses import Response
    filename = Path(remote).name
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@app.get("/memory")
def list_memory(_: str = Depends(require_auth)):
    memory = load_memory()
    return [{"key":k,**v} for k,v in memory.items()]

app.mount("/",StaticFiles(directory=str(BASE_DIR/"static"),html=True),name="static")
