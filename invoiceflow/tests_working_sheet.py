"""Working-sheet layout tests — the commodity code as two customs columns.

The sheet the reviewer reads shows the declared code split the way customs
reads it: 8 digits of nomenclature in "Comm./imp. cod", the additional
2-digit code in "TARIC". Adding that column also shifts the title merge,
the totals merge and the SUM columns, so those are asserted here too.

Imports main, so (like tests_user_admin.py) it relies on a valid .env.
Run: cd invoiceflow && python -m pytest tests_working_sheet.py -q
"""
import io
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent))
import main as m  # noqa: E402 — relies on .env being valid


def _row(code, desc="GOODS", **over):
    row = {"Invoice": "610", "Comm./imp. cod": code, "Description of Goods": desc,
           "Origin": "IT", "Country": "Italy", "Number of Packages": "10",
           "Gross Weight (KG)": "120.5", "Net Weight (KG)": "110.25",
           "Value": "€500.00"}
    row.update(over)
    return row


def _sheet(rows, **kw):
    data = m.build_excel(rows, None, "Test", **kw)
    return openpyxl.load_workbook(io.BytesIO(data)).active


def _headers(ws):
    return [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]


def test_taric_is_its_own_column_right_after_the_code():
    ws = _sheet([_row("2005998098")])
    assert _headers(ws)[:3] == ["Invoice", "Comm./imp. cod", "TARIC"]


def test_a_10_digit_code_is_written_as_8_plus_2():
    ws = _sheet([_row("2005998098")])
    assert ws.cell(row=3, column=2).value == "20059980"
    assert ws.cell(row=3, column=3).value == "98"


def test_an_8_digit_code_leaves_the_taric_cell_empty():
    ws = _sheet([_row("07049010")])
    assert ws.cell(row=3, column=2).value == "07049010"
    assert ws.cell(row=3, column=3).value in (None, "")


def test_both_parts_keep_their_leading_zeros_as_text():
    ws = _sheet([_row("0704901000")])
    assert ws.cell(row=3, column=2).value == "07049010"
    assert ws.cell(row=3, column=3).value == "00"
    assert ws.cell(row=3, column=2).number_format == "@"
    assert ws.cell(row=3, column=3).number_format == "@"


def test_a_value_that_is_not_a_code_is_shown_untouched():
    # An internal SKU must not be sliced into something that reads as a code.
    ws = _sheet([_row("22.289", "INTERNAL SKU")])
    assert ws.cell(row=3, column=2).value == "22.289"
    assert ws.cell(row=3, column=3).value in (None, "")


def test_a_repaired_leading_zero_is_flagged_not_silent():
    ws = _sheet([_row("4061030", "CHEESE")])
    assert ws.cell(row=3, column=2).value == "04061030"
    flag = m._FILL_FLAG.fgColor.rgb
    assert ws.cell(row=3, column=2).fill.fgColor.rgb == flag
    assert ws.cell(row=3, column=3).fill.fgColor.rgb == flag


def test_a_disagreement_on_the_code_highlights_both_parts():
    ws = _sheet([_row("2005998098")], flagged_cells=[{"Comm./imp. cod"}])
    flag = m._FILL_FLAG.fgColor.rgb
    assert ws.cell(row=3, column=2).fill.fgColor.rgb == flag
    assert ws.cell(row=3, column=3).fill.fgColor.rgb == flag


def test_a_clean_code_is_not_highlighted():
    ws = _sheet([_row("2005998098")])
    flag = m._FILL_FLAG.fgColor.rgb
    assert ws.cell(row=3, column=2).fill.fgColor.rgb != flag


# ── The extra column must not leave the sheet furniture behind ──────────
def test_title_and_totals_span_the_widened_sheet():
    ws = _sheet([_row("2005998098"), _row("0704901000")])
    merges = {str(r) for r in ws.merged_cells.ranges}
    assert f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}1" in merges
    total_row = 3 + 2                       # two data rows after the header
    assert f"A{total_row}:F{total_row}" in merges   # up to the last text column
    assert ws.cell(row=total_row, column=1).value == "TOTALS"


def test_totals_still_sum_the_numeric_columns():
    ws = _sheet([_row("2005998098"), _row("0704901000")])
    total_row = 5
    sums = [ws.cell(row=total_row, column=c).value for c in range(7, 11)]
    assert sums == [f"=SUM({col}3:{col}4)" for col in ("G", "H", "I", "J")]
