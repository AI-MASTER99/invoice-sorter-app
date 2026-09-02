"""MultiFreight Items export — the DE 2/1 previous-document columns.

Every goods line the export writes refers back to the commercial invoice
it was declared from: category Z, type 380, reference = the invoice
number as printed on the invoice (never a customer code or order number).
The invoice number is the one the extraction step read off the invoice,
so a wrong reading shows up in this column — the reference is left blank
rather than guessed when no number is known.

Imports main, so it relies on a valid .env (dummy values are fine).
Run: cd invoiceflow && python -m pytest tests_items_export.py -q
"""
import io
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import main as m  # noqa: E402 — relies on .env being valid


def _row(code="20060035", invoice="0002236", **over):
    row = {"Invoice": invoice, "Comm./imp. cod": code,
           "Description of Goods": "GOODS", "Origin": "IT", "Country": "Italy",
           "Number of Packages": "10", "Gross Weight (KG)": "120.5",
           "Net Weight (KG)": "110.25", "Value": "€500.00"}
    row.update(over)
    return row


def _items(rows, totals=None):
    data = m.build_items_xlsx(rows, totals=totals)
    ws = openpyxl.load_workbook(io.BytesIO(data))["Items"]
    col = {str(ws.cell(row=3, column=c).value).strip(): c
           for c in range(1, ws.max_column + 1)
           if ws.cell(row=3, column=c).value not in (None, "")}
    return ws, col


def _cell(ws, col, r, header):
    return ws.cell(row=r, column=col[header])


def test_every_line_refers_back_to_the_invoice_as_z_380():
    ws, col = _items([_row("20060035"), _row("08042090")])
    for r in (4, 5):
        assert _cell(ws, col, r, "[2/1] Previous Documents - Category (01)").value == "Z"
        assert _cell(ws, col, r, "[2/1] Previous Documents - Type (01)").value == "380"
        assert _cell(ws, col, r, "[2/1] Previous Documents - Ref (01)").value == "0002236"


def test_previous_document_cells_are_text_so_leading_zeros_survive():
    ws, col = _items([_row()])
    for h in ("[2/1] Previous Documents - Category (01)",
              "[2/1] Previous Documents - Type (01)",
              "[2/1] Previous Documents - Ref (01)"):
        assert _cell(ws, col, 4, h).number_format == "@"


def test_the_reference_is_the_invoice_number_not_anything_else():
    # The row's Invoice column is what the extraction read as the invoice
    # number; nothing else on the row (code, origin, REX) may leak into it.
    ws, col = _items([_row(invoice="FT-2026/118")],
                     totals={"supplier_rex": "ITREXIT06678581213"})
    assert _cell(ws, col, 4, "[2/1] Previous Documents - Ref (01)").value == "FT-2026/118"
    # And the N935 invoice document carries the same number.
    assert _cell(ws, col, 4, "[2/3] Documents - ID (01)").value == "VM1 FT-2026/118"


def test_an_unknown_invoice_number_leaves_the_reference_blank_not_guessed():
    ws, col = _items([_row(invoice="")])
    assert _cell(ws, col, 4, "[2/1] Previous Documents - Category (01)").value == "Z"
    assert _cell(ws, col, 4, "[2/1] Previous Documents - Type (01)").value == "380"
    assert _cell(ws, col, 4, "[2/1] Previous Documents - Ref (01)").value in (None, "")


def test_only_the_first_previous_document_slot_is_used():
    ws, col = _items([_row()])
    for h in ("[2/1] Previous Documents - Category (02)",
              "[2/1] Previous Documents - Type (02)",
              "[2/1] Previous Documents - Ref (02)",
              "[2/1] Previous Documents - Line (01)"):
        assert _cell(ws, col, 4, h).value in (None, "")


def test_a_fee_row_is_not_an_items_line():
    ws, col = _items([_row(), _row(code="", **{"Description of Goods": "TRANSPORT"})])
    assert _cell(ws, col, 5, "[2/1] Previous Documents - Category (01)").value in (None, "")
